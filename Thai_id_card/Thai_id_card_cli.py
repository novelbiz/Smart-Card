
# -*- Thai ID Smart Card Reader with Face Recognition -*- 

from PyQt5.QtWidgets import (QApplication, QLabel, QPushButton, QMainWindow, 
                             QMessageBox, QCheckBox, QLineEdit, QGroupBox, 
                             QVBoxLayout, QHBoxLayout, QWidget, QFrame, QGridLayout,
                             QFileDialog, QProgressBar, QGraphicsDropShadowEffect)
from PyQt5.QtCore import QTimer, Qt, QLocale, QThread, pyqtSignal, QObject
from PyQt5.QtGui import QPixmap, QFont, QColor, QIcon, QImage ,QPalette
from smartcard.System import readers
from smartcard.CardMonitoring import CardMonitor, CardObserver
from smartcard.CardType import AnyCardType
from smartcard.CardRequest import CardRequest
from smartcard.Exceptions import NoCardException
from smartcard.util import toHexString
from smartcard.scard import SCARD_PROTOCOL_T0, SCARD_PROTOCOL_T1, SCARD_SHARE_SHARED
from deepface import DeepFace
import tempfile
import subprocess
import os
import io
from datetime import datetime
import numpy as np
import cv2
from PIL import ImageFont, ImageDraw, Image
import sqlite3
from database_viewer import DatabaseViewerWindow

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.drawing.image import Image as OpenpyxlImage
    from openpyxl.styles import Font, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

class IDCardObserver(CardObserver):
    def __init__(self, window):
        self.window = window

    def update(self, observable, actions):
        (addedcards, removedcards) = actions
        if addedcards:
            self.window.on_card_inserted()
        if removedcards:
            self.window.on_card_removed()


class DeepFaceWorker(QThread):
    """Worker thread สำหรับประมวลผล DeepFace แบบไม่บล็อก UI"""
    result_signal = pyqtSignal(bool, float)  # verified, distance
    error_signal = pyqtSignal(str)  # error_message
    
    def __init__(self, card_photo_bytes, captured_frame, parent=None):
        super().__init__(parent)
        self.card_photo_bytes = card_photo_bytes
        self.captured_frame = captured_frame
        
    def run(self):
        """ทำงานใน background thread"""
        try:

            
            # บันทึกรูปชั่วคราว
            with tempfile.NamedTemporaryFile(suffix='.jpg', delete=False) as f:
                f.write(self.card_photo_bytes)
                card_path = f.name
                
            with tempfile.NamedTemporaryFile(suffix='.jpg', delete=False) as f:
                cv2.imwrite(f.name, self.captured_frame)
                captured_path = f.name
            
            try:
                # เรียก DeepFace
                result = DeepFace.verify(
                    img1_path=card_path,
                    img2_path=captured_path,
                    model_name="Facenet512",
                    detector_backend="opencv",
                    enforce_detection=False,
                    align=True
                )
                
                verified = result["verified"]
                distance = result["distance"]
                
                # ส่งผลลัพธ์กลับ
                self.result_signal.emit(verified, distance)
                
            finally:
                # ลบไฟล์ชั่วคราว
                try:
                    os.unlink(card_path)
                    os.unlink(captured_path)
                except:
                    pass
                    
        except Exception as e:
            self.error_signal.emit(str(e))
            
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.base_width = 870
        self.base_height = 870
        self.base_font_size = 13
        self.ui_ready = False 
        
        self.camera = None 
        self.webcam_label = None # จะถูกสร้างใน setup_ui()
        self.is_face_scan_mode = False 
        self.captured_face_frame = None 
        self.current_data = {}
        self.cardservice = None
        self.current_reader = None

        self.face_cascade = cv2.CascadeClassifier(
            cv2.data.haarcascades + 'haarcascade_frontalface_default.xml') 
        
        # 📌📌 โค้ดที่คุณต้องเพิ่มใน __init__ เพื่อยืนยันการโหลด
        if self.face_cascade.empty():
            print("FATAL ERROR: Haar Cascade XML file not found or failed to load!")
            self.face_cascade = None # ตั้งค่าให้เป็น None เพื่อให้โค้ด update_frame ตรวจจับได้

         
        # Timer สำหรับ WebCam (ใช้สำหรับดึงเฟรม)
        self.webcam_timer = QTimer(self)
        self.webcam_timer.timeout.connect(self.update_frame) # 📌 เชื่อมต่อสัญญาณตรงนี้
        
        # Timer สำหรับ Card Reader Status
        self.timer = QTimer()
        self.timer.timeout.connect(self.update_reader_status)
        self.timer.start(2000) # เริ่ม Timer ตรวจสอบสถานะ Card Reader
        
        documents_path = os.path.join(os.path.expanduser('~'), 'Documents')
        target_path = os.path.join(documents_path, 'Novelbiz')
        self.folder_path = QLineEdit(target_path)
        
        main_data_folder = os.path.join(self.folder_path.text(), "ID_Card_Data")
        if not os.path.exists(main_data_folder):
            os.makedirs(main_data_folder, exist_ok=True)
            
        self.db_name = os.path.join(main_data_folder, 'id_card_data.db')
        self.init_db() 
                
        icon_path = os.path.join(os.path.dirname(__file__), "logo.png")
        self.setWindowIcon(QIcon(icon_path))
        self.setWindowTitle("Thai_ID_Card_Reader.v1.4_FaceRec")
        
        self.setGeometry(100, 50, 900, 900)        
        
        # สำหรับ MainWindow + QPushButton
        self.setStyleSheet("""
            QMainWindow {
                background-color: #f0f4f8;
            }
          
            QGroupBox {
                background: #ffffff;
                border: 1px solid #e5e7eb;
                border-radius: 12px;
                margin-top: 16px;
                padding: 16px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 6px;
                font-weight: bold;
                color: #1e40af;
                font-size: 16px;
            }
            QPushButton {
            background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                        stop:0 #60a5fa, stop:0.5 #3b82f6, stop:1 #2563eb);
                            border: 1px solid #1e40af;
                            border-radius: 8px;
                            color: white;
                            font-size: 13px;
                            font-weight: bold;
                            padding: 6px 14px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                            stop:0 #2563eb, stop:1 #1d4ed8);
            }
            QPushButton:pressed {
                background-color: #1e3a8a;
            }
            QPushButton:disabled {
                background: #cbd5e0;
                color: #94a3b8;
            }
        """)
        
      
        self.setup_ui()  
        self.start_webcam()            
        self.update_reader_status()            
        # Card monitoring
        self.cardmonitor = CardMonitor()
        self.cardobserver = IDCardObserver(self)
        self.cardmonitor.addObserver(self.cardobserver)

     

    # ------------------- UI Setup -------------------
    def setup_ui(self):
        # ใช้ layout แบบเต็มหน้าต่าง ไม่ต้อง centering
        central_widget = QWidget()
        self.setCentralWidget(central_widget)    
        # Main layout
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(8)
        main_layout.setContentsMargins(15, 15, 15, 15)
        # Status frame
        status_frame = QFrame(self)
        status_frame.setFrameShape(QFrame.Box)
        status_frame.setStyleSheet("""
            QFrame {
                background-color: #f0f4f8;
                border-radius: 8px;
                padding: 4px;
            }
        """)
        
        status_layout = QHBoxLayout(status_frame)
        status_layout.setContentsMargins(2, 2, 2, 2)
        status_layout.setSpacing(5)
        font_small = QFont("Prompt", 9)  # ลดจาก 10
        font_small.setBold(True)

        # Reader Status
        self.reader_label = QLabel("สถานะเครื่องอ่าน: กำลังรอ...")
        self.reader_label.setFont(font_small)
        self.reader_label.setStyleSheet("color: blue;")

        # Card Status
        self.card_label = QLabel("สถานะบัตรประชาชน: ยังไม่มีบัตร")
        self.card_label.setFont(font_small)
        self.card_label.setStyleSheet("color: red;")
        
        status_layout.addStretch()
        status_layout.addWidget(self.reader_label)
        status_layout.addWidget(self.card_label)
        status_layout.addStretch()       
        main_layout.addWidget(status_frame)    
          
        GROUP_HEIGHT = 210 
        # Data Layout
        data_layout = QHBoxLayout()
        data_layout.setSpacing(15)
        
        # Left: Personal Info
        left_layout = QVBoxLayout()
        group_font = QFont()
        group_font.setFamily("Prompt")
        group_font.setPointSize(10)  # ลดขนาด font
        group_font.setBold(True)

        # Personal Info Group (ลดความสูง)
        personal_group = QGroupBox("ข้อมูลบัตรประจำตัวประชาชน")
        personal_group.setMinimumHeight(GROUP_HEIGHT)
        personal_group.setFont(group_font)
        personal_group.setStyleSheet("""
            QGroupBox {
                font-family: "Prompt";
                border: 1px solid #a6a6a6;
                border-radius: 8px;
                margin-top: 16px;
                padding: 12px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                subcontrol-position: top left;
                padding: 2px 5px;
                font-weight: bold;
                color: #3B82F6;
            }
            QLabel {
                font-family: "Prompt";
                font-size: 13px;
                line-height: 10px;
                margin: 0px;
                padding: 0px;
                color: #333333;
            }
        """)
        personal_layout = QVBoxLayout(personal_group)
        personal_layout.setSpacing(3)  # ลด spacing

        # CID & Gender
        self.cid_label = QLabel("เลขประจำตัวประชาชน :")
        personal_layout.addWidget(self.cid_label)
        self.gender_label = QLabel("เพศ :")
        personal_layout.addWidget(self.gender_label)

        # Name
        self.name_th_label = QLabel("ชื่อ - นามสกุล :")
        personal_layout.addWidget(self.name_th_label)
        self.name_en_label = QLabel("ชื่อ - นามสกุล (Eng) :")
        personal_layout.addWidget(self.name_en_label)

        # Birth
        self.birth_th_label = QLabel("วันเกิด (ไทย) :")
        personal_layout.addWidget(self.birth_th_label)
        self.birth_en_label = QLabel("วันเกิด (Eng) :") 
        personal_layout.addWidget(self.birth_en_label)

        # Address
        self.address_label = QLabel("ที่อยู่ :")
        self.address_label.setWordWrap(True)
        personal_layout.addWidget(self.address_label)
        personal_layout.addStretch()
        left_layout.addWidget(personal_group, 1)

        # Card Info Group (ปรับให้กะทัดรัด)
        address_group = QGroupBox("ข้อมูลการออกบัตร")
        address_group.setMinimumHeight(GROUP_HEIGHT)
        address_group.setFont(group_font)
        address_group.setStyleSheet("""
            QGroupBox {
                font-family: "Prompt";
                border: 1px solid #a6a6a6;
                border-radius: 8px;
                margin-top: 16px;
                padding: 12px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                subcontrol-position: top left;
                padding: 2px 5px;
                font-weight: bold;
                color: #3B82F6;
            }
            QLabel {
                font-family: "Prompt";
                font-size: 13px;
                line-height: 10px;
                margin: 0px;
                padding: 0px;
                color: #333333;
            }
        """)

        address_layout = QGridLayout(address_group)
        address_layout.setSpacing(1)
        address_layout.setContentsMargins(6, 6, 6, 6)

        self.issue_date_label = QLabel("วันที่ออกบัตร :")
        self.issue_date_en_label = QLabel("วันที่ออกบัตร (Eng) :")
        self.expire_date_label = QLabel("วันที่หมดอายุ :")
        self.expire_date_en_label = QLabel("วันที่หมดอายุ (Eng) :")
        self.issuer_label = QLabel("สถานที่ออกบัตร :")
        self.issuer_label.setWordWrap(True)

        address_layout.addWidget(self.issue_date_label, 0, 0)
        address_layout.addWidget(self.issue_date_en_label, 1, 0)
        address_layout.addWidget(self.expire_date_label, 2, 0)
        address_layout.addWidget(self.expire_date_en_label, 3, 0)
        address_layout.addWidget(self.issuer_label, 4, 0, 1, 1)
        address_layout.setRowStretch(5, 1)
        
        left_layout.addWidget(address_group, 1)

        # Right: Photo & Webcam (ลดขนาด)
        right_layout = QVBoxLayout()
        right_layout.setContentsMargins(0, 10, 0, 0)
        right_layout.setSpacing(8)
        
        # รูปบัตร (ลดขนาด)
        self.photo_label = QLabel()
        self.photo_label.setMinimumSize(240, 200)  # ลดจาก 300x240
        self.photo_label.setStyleSheet("""
            border: 2px dashed #a6a6a6;
            background-color: #ffffff;
            border-radius: 8px;
        """)
        self.photo_label.setAlignment(Qt.AlignCenter)
        self.photo_label.setText("รูปภาพจากบัตรประชาชน")
        right_layout.addWidget(self.photo_label, 1)
        
        # หมายเลขคำขอ
        self.request_number_label = QLabel("หมายเลขคำขอการมีบัตร :")
        self.request_number_label.setAlignment(Qt.AlignCenter)
        self.request_number_label.setStyleSheet("font-size: 13px; font-weight: normal; color: #555555;")
        right_layout.addWidget(self.request_number_label)
                    
        # กล้องเว็บแคม (ลดขนาด)
        self.webcam_label = QLabel()
        self.webcam_label.setMinimumSize(240, 170)  # ลดจาก 300x220
        self.webcam_label.setStyleSheet("""
            border: 2px solid #3B82F6;
            background-color: #f0f4f8;
            border-radius: 8px;
        """)
        self.webcam_label.setAlignment(Qt.AlignCenter)
        self.webcam_label.setText("กล้องเว็บแคม")
        right_layout.addWidget(self.webcam_label, 1)
        
        right_layout.addStretch()

        data_layout.addLayout(left_layout, 3)
        data_layout.addLayout(right_layout, 2)
        
        main_layout.addLayout(data_layout)

        # Log Group (ลดความสูง)
        log_group = QGroupBox("Smart Card Data Options")
        log_group.setFont(group_font)
        log_group.setStyleSheet("""
            QGroupBox {
                font-family: "Prompt";
                border: 1px solid #a6a6a6;
                border-radius: 8px;
                margin-top: 16px;
                padding: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                subcontrol-position: top left;
                padding: 2px 5px;
                font-weight: bold;
                color: #3B82F6;
            }
            QLabel {
                font-family: "Prompt";
                font-size: 13px;
                color: #333333;
            }
        """)
        log_layout = QVBoxLayout(log_group)
        log_layout.setSpacing(5)

        # Folder selection
        folder_layout = QHBoxLayout()
        folder_layout.setSpacing(8)
        folder_layout.addWidget(QLabel("Folder :"))
        
        documents_path = os.path.join(os.path.expanduser('~'), 'Documents')
        target_path = os.path.join(documents_path, 'Novelbiz')
        self.folder_path = QLineEdit(target_path)
        
        self.folder_path.setStyleSheet("""
            QLineEdit {
                background: #ffffff;
                border: 1px solid #cbd5e0;
                border-radius: 6px;
                padding: 4px 10px;
                font-size: 12px;
                color: #1f2937;
            }
            QLineEdit:focus {
                border: 1px solid #3b82f6;
                background: #f0f9ff;
            }
        """)
        folder_layout.addWidget(self.folder_path)
        
        browse_btn = QPushButton("Browse")
        browse_btn.setFixedSize(80, 28)
        browse_btn.clicked.connect(self.browse_folder)
        folder_layout.addWidget(browse_btn)
        log_layout.addLayout(folder_layout)

        # Data type checkboxes
        data_layout_cb = QHBoxLayout()
        data_layout_cb.addWidget(QLabel("Data :"))
        self.text_file_cb = QCheckBox("Text File"); self.text_file_cb.setChecked(True)
        self.excel_file_cb = QCheckBox("Excel File"); self.excel_file_cb.setChecked(True)
        self.image_cb = QCheckBox("Image"); self.image_cb.setChecked(True)
        self.sqlite_cb = QCheckBox("DB"); self.sqlite_cb.setChecked(True)
        for cb in [self.text_file_cb, self.excel_file_cb, self.image_cb, self.sqlite_cb]:
            cb.setStyleSheet("font-family: 'Prompt'; font-size: 12px; color: #333333;")
        data_layout_cb.addWidget(self.text_file_cb)
        data_layout_cb.addWidget(self.excel_file_cb)
        data_layout_cb.addWidget(self.image_cb)
        data_layout_cb.addWidget(self.sqlite_cb)
        log_layout.addLayout(data_layout_cb)

        log_layout_container = QHBoxLayout()
        log_layout_container.addWidget(log_group, 3)  # 3 = สัดส่วนเท่า left_layout
        # เพิ่ม QGroupBox ใหม่ต่อท้าย log_group
        
        new_group = QGroupBox("สถานะยืนยันใบหน้า")
        new_group.setFont(group_font)
        new_group.setStyleSheet("""
            QGroupBox {
                font-family: "Prompt";
                border: 1px solid #a6a6a6;
                border-radius: 8px;
                margin-top: 16px;
                padding: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                padding: 2px 5px;
                font-weight: bold;
                color: #3B82F6;
            }
        """)
        new_layout = QVBoxLayout(new_group)
        # เพิ่ม Label แสดงสถานะ
        self.face_verify_status = QLabel("รอการตรวจสอบ")
        self.face_verify_status.setAlignment(Qt.AlignCenter)
        self.face_verify_status.setStyleSheet("""
            font-size: 14px;
            font-weight: bold;
            color: #6B7280;
            padding: 10px;
        """)
        new_layout.addWidget(self.face_verify_status)
        # เพิ่ม Progress Bar
        self.face_progress = QProgressBar()
        self.face_progress.setVisible(False)  # ซ่อนไว้ก่อน
        self.face_progress.setStyleSheet("""
            QProgressBar {
                border: 2px solid #3B82F6;
                border-radius: 5px;
                text-align: center;
                font-size: 12px;
            }
            QProgressBar::chunk {
                background-color: #3B82F6;
            }
        """)
        self.face_progress.setFormat("%p%") 
        self.face_progress.setLocale(QLocale.c()) 
        new_layout.addWidget(self.face_progress)        
        log_layout_container.addWidget(new_group, 2)  # 2 = เว้นที่ว่างด้านขวา (เท่า right_layout)
        main_layout.addLayout(log_layout_container)
        


        # Buttons (ลดขนาด)
        button_layout = QHBoxLayout()
        button_layout.setSpacing(8)
        button_layout.addStretch()
        self.read_button = QPushButton("Read ID Card")
        self.read_button.setEnabled(False)
        self.read_button.clicked.connect(self.read_id_card)

        # <<< โค้ดที่เพิ่ม: ปุ่มบันทึกข้อมูล >>>
        self.save_button = QPushButton("Save") # สร้างปุ่ม
        self.save_button.setEnabled(False) # ปิดใช้งานปุ่มก่อนอ่านข้อมูล
        self.save_button.clicked.connect(self.save_data) # เชื่อมต่อกับฟังก์ชัน self.save_data        
        
        self.scan_face_button = QPushButton("Face scan")
        self.scan_face_button.setEnabled(False)
        self.scan_face_button.clicked.connect(self.scan_and_compare_face)
        
        self.clear_button = QPushButton("Clear")
        self.clear_button.clicked.connect(self.clear_data)
        
        self.exit_button = QPushButton("Exit")
        self.exit_button.clicked.connect(self.exit_application)

        # เพิ่มปุ่มดูฐานข้อมูล
        self.view_db_button = QPushButton("ดูฐานข้อมูล")
        self.view_db_button.clicked.connect(self.open_database_viewer)
        
        # แก้ไข loop ปุ่ม
        for btn in [self.read_button, self.scan_face_button, self.save_button, 
                    self.view_db_button, self.clear_button, self.exit_button]:
            btn.setFixedSize(120, 35)
            button_layout.addWidget(btn)

        button_layout.addStretch()
        


        bottom_section_layout = QHBoxLayout()
        bottom_section_layout.addLayout(button_layout)
        
        main_layout.addLayout(bottom_section_layout)
        
        self.ui_ready = True
        self.update_fonts(self.base_font_size)
        
    # ------------------- Helper -------------------
    def set_label_style(self, labels, font_size=12, bold=False):
        style = f"font-size:{font_size}px;{' font-weight:bold;' if bold else ''}; color:#222222;"
        for lbl in labels:
            lbl.setStyleSheet(style)

    def browse_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "Select Folder")
        if folder:
            self.folder_path.setText(folder)

    # ------------------- Service & Reader -------------------
    def start_service(self):
        try:
            # สั่ง start service
            subprocess.run(["sc", "start", "SCardSvr"], capture_output=True, text=True, shell=True)
            QMessageBox.information(self, "Service", "Smart Card Service started (if available).")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Cannot start service: {e}")
           
    def check_service_status(self):
        try:
            result = subprocess.run(
                ["sc", "query", "SCardSvr"],
                capture_output=True, text=True, shell=True
            )
            return "RUNNING" in result.stdout
        except Exception as e:
            print(f"Service check error: {e}")
            return False

    def update_reader_status(self):
        if not self.check_service_status():
            self.reader_label.setText("สถานะเครื่องอ่าน: บริการ Smart Card ไม่ทำงาน")
            self.reader_label.setStyleSheet("font-weight: bold; color: orange;")
            self.card_label.setText("สถานะบัตรประชาชน: บริการไม่ทำงาน")
            self.card_label.setStyleSheet("font-weight: bold; color: red;")
            self.read_button.setEnabled(False)
            return

        r = readers()
        if r:
            self.current_reader = r[0]
            self.reader_label.setText(f"สถานะเครื่องอ่าน: เชื่อมต่อแล้ว")
            self.reader_label.setStyleSheet("font-weight: bold; color: green;")
        else:
            self.current_reader = None
            self.reader_label.setText("สถานะเครื่องอ่าน: ไม่พบเครื่องอ่าน")
            self.reader_label.setStyleSheet("font-weight: bold; color: red;")
            self.card_label.setText("สถานะบัตรประชาชน: ยังไม่มีบัตร")
            self.card_label.setStyleSheet("font-weight: bold; color: red;")
            self.read_button.setEnabled(False)

    # ------------------- Card Insert/Remove -------------------
    def on_card_inserted(self):
        self.card_label.setText("สถานะบัตรประชาชน: ใส่บัตรแล้ว")
        self.card_label.setStyleSheet("font-weight: bold; color: green;")
        self.read_button.setEnabled(True)            
        
            
    def on_card_removed(self):
        self.card_label.setText("สถานะบัตรประชาชน: ยังไม่มีบัตร")
        self.card_label.setStyleSheet("font-weight: bold; color: red;")
        self.read_button.setEnabled(False)
        self.scan_face_button.setEnabled(False)
        self.clear_data()

    # ------------------- Decode & Parse -------------------
    def decode_text(self, data):
        try:
            return bytes(data).decode('tis-620', errors='ignore').strip()
        except:
            return ''.join(chr(b) if b < 128 else '?' for b in data).strip()

    def send_apdu_with_get_response(self, connection, apdu):
        """Send APDU and handle GET_RESPONSE if needed"""
        response, sw1, sw2 = connection.transmit(apdu)
        if sw1 == 0x61:
            get_response = [0x00, 0xC0, 0x00, 0x00, sw2]
            response, sw1, sw2 = connection.transmit(get_response)
        return response, sw1, sw2

    def parse_thai_date(self, date_str):
            """Parse Thai date format (YYYYMMDD) and convert to readable format"""
            # กรณีที่วันหมดอายุเป็น 99999999 ให้แสดงเป็น ตลอดชีพ / LIFELONG
            if date_str == '99999999':
                return "ตลอดชีพ", "LIFELONG"
                
            # ตรวจสอบว่าสตริงวันที่มีความยาว 8 ตัวอักษรและเป็นตัวเลขเท่านั้น
            if len(date_str) == 8 and date_str.isdigit():
                try:
                    year = date_str[0:4]
                    month = date_str[4:6]
                    day = date_str[6:8]
                    
                    # Convert Buddhist year to Christian year for English
                    thai_year = int(year)
                    eng_year = thai_year - 543
                    
                    # Thai month names
                    thai_months = ['', 'มกราคม', 'กุมภาพันธ์', 'มีนาคม', 'เมษายน', 'พฤษภาคม', 'มิถุนายน',
                                'กรกฎาคม', 'สิงหาคม', 'กันยายน', 'ตุลาคม', 'พฤศจิกายน', 'ธันวาคม']
                    
                    eng_months = ['', 'January', 'February', 'March', 'April', 'May', 'June',
                                'July', 'August', 'September', 'October', 'November', 'December']
                    
                    month_int = int(month)
                    if 1 <= month_int <= 12:
                        thai_date = f"{int(day)} {thai_months[month_int]} {thai_year}"
                        eng_date = f"{int(day)} {eng_months[month_int]} {eng_year}"
                        return thai_date, eng_date
                except Exception as e:
                    print(f"Date parsing error: {e}")
                    return "ไม่ระบุ", "Not specified"
            else:
                # หากสตริงไม่ตรงตามรูปแบบที่คาดหวัง ให้คืนค่าเป็น "ไม่ระบุ"
                return "ไม่ระบุ", "Not specified"          

    def disconnect_card(self):
        """Safely disconnect from card"""
        if self.cardservice:
            try:
                self.cardservice.connection.disconnect()
                print("Card disconnected successfully")
            except Exception as e:
                print(f"Error disconnecting card: {e}")
            finally:
                self.cardservice = None

    # ------------------- Face Recognition -------------------
    def compare_faces(self, captured_frame, face_cascade):
        """เริ่มการประมวลผลเปรียบเทียบใบหน้าด้วย DeepFace ใน Worker Thread"""
        
        # ✅ ตรวจสอบข้อมูลครบถ้วนก่อนดำเนินการ
        if not self.current_data:
            QMessageBox.warning(self, "ข้อผิดพลาด", "ไม่พบข้อมูลบัตรประชาชน\nกรุณาอ่านบัตรก่อน")
            return
            
        if 'photo' not in self.current_data:
            QMessageBox.warning(self, "ข้อผิดพลาด", "ไม่พบรูปภาพในบัตร\nกรุณาอ่านบัตรใหม่อีกครั้ง")
            return
        
        if self.current_data['photo'] is None or len(self.current_data['photo']) == 0:
            QMessageBox.warning(self, "ข้อผิดพลาด", "รูปภาพในบัตรไม่สมบูรณ์\nกรุณาอ่านบัตรใหม่อีกครั้ง")
            return

        # แสดง progress bar
        self.face_progress.setVisible(True)
        self.face_progress.setValue(0)
        
        # อัปเดตสถานะ GUI
        self.face_verify_status.setText("กำลังตรวจสอบใบหน้าด้วย DeepFace... (โปรดรอ)")
        self.face_verify_status.setStyleSheet(
            "font-size: 14px; font-weight: bold; color: #F59E0B; padding: 10px;"
        )
        
        try:
            # สร้าง Worker
            card_photo_bytes = self.current_data['photo']
            success, encoded_image = cv2.imencode('.jpg', captured_frame)
            if success:
                captured_photo_bytes = encoded_image.tobytes()
                # เก็บรูปที่ถ่ายจากกล้องไว้ใน current_data
                self.current_data['captured_face_photo'] = captured_photo_bytes
                print("✅ เก็บรูปภาพจากกล้องสำเร็จ")
            else:
                print("⚠️ ไม่สามารถแปลงรูปภาพจากกล้องได้")        
            
            
            self.face_worker = DeepFaceWorker(
                card_photo_bytes, 
                captured_frame,
                parent=self
            )
            
            # เชื่อมต่อ Signals
            self.face_worker.result_signal.connect(self.handle_face_result)
            self.face_worker.error_signal.connect(self.handle_face_error)
            
            # เริ่มทำงาน
            self.face_worker.start()
            
        except Exception as e:
            self.face_progress.setVisible(False)
            self.face_verify_status.setText("ข้อผิดพลาดในการตรวจสอบ")
            self.face_verify_status.setStyleSheet(
                "font-size: 14px; font-weight: bold; color: #EF4444; padding: 10px;"
            )
            QMessageBox.critical(self, "ข้อผิดพลาด", f"ไม่สามารถเริ่มการตรวจสอบได้:\n{str(e)}")
            
    def handle_face_result(self, verified, distance):
        """รับผลลัพธ์จาก Worker (DeepFace) และอัปเดต GUI + แสดง QMessageBox"""
        
        cid = self.current_data.get('cid', 'ไม่ระบุ')
        name_th = self.current_data.get('name_th', 'ไม่ระบุ')
        
        # คำนวณความคล้ายคลึง
        similarity_percent = max(0, (1 - distance) * 100) 
        
        if verified:
            # ✅ ตรงกัน
            self.face_verify_status.setText("ผลการตรวจสอบ: ตรงกับบัตรประชาชน")
            self.face_verify_status.setStyleSheet(
                "font-size: 14px; font-weight: bold; color: #10B981; padding: 10px;"
            )
                    # ✅ เปิดปุ่ม Save ได้แล้ว
            self.save_button.setEnabled(True)
            print("✅ ยืนยันใบหน้าสำเร็จ - เปิดปุ่ม Save แล้ว")
            QMessageBox.information(
                self, "ผลการตรวจสอบ",
                f"<h3 style='color: green;'>✔ ใบหน้าตรงกับบัตรประชาชน</h3>"
                f"<b>ความเหมือน:</b> {similarity_percent:.1f}%<br>"
                f"<b>Distance:</b> {distance:.4f}<br><br>"
                f"<b>เลขบัตร:</b> {cid}<br>"
                f"<b>ชื่อ:</b> {name_th}"
            )
        else:
            # ❌ ไม่ตรงกัน
            self.face_verify_status.setText("ผลการตรวจสอบ: ไม่ตรงกับบัตรประชาชน")
            self.face_verify_status.setStyleSheet(
                "font-size: 14px; font-weight: bold; color: #EF4444; padding: 10px;"
            )
            self.save_button.setEnabled(True)
            print("❌ ใบหน้าไม่ตรง - ปิดปุ่ม Save")            
            QMessageBox.warning(
                self, "ผลการตรวจสอบ",
                f"<h3 style='color: red;'>✗ ใบหน้าไม่ตรงกับบัตรประชาชน</h3>"
                f"<b>ความเหมือน:</b> {similarity_percent:.1f}%<br>"
                f"<b>Distance:</b> {distance:.4f}<br><br>"
                f"<b>เลขบัตร:</b> {cid}<br>"
                f"<b>ชื่อ:</b> {name_th}<br><br>"
                f"<small>โปรดลองสแกนใหม่ให้ใบหน้าตรงกับกล้อง</small>"
            )
            
        # ซ่อน progress bar
        self.face_progress.setVisible(False)
            
    def handle_face_error(self, error_message):
        """รับข้อผิดพลาดจาก Worker"""
        self.face_verify_status.setText("ข้อผิดพลาดในการตรวจสอบ")
        self.face_verify_status.setStyleSheet(
            "font-size: 14px; font-weight: bold; color: #9D174D; padding: 10px;"
        )
        self.face_progress.setVisible(False)
        
        QMessageBox.critical(
            self, "ข้อผิดพลาด DeepFace", 
            f"เกิดข้อผิดพลาดในการประมวลผล:\n{error_message}"
        )
    # ------------------- Read ID Card -------------------
    def read_id_card(self):
        try:
            # Disconnect any existing connection first
            self.disconnect_card()
            
            # Connect to card
            cardtype = AnyCardType()
            cardrequest = CardRequest(timeout=3, cardType=cardtype)
            self.cardservice = cardrequest.waitforcard()
            
            print(f'CONNECTED TO {self.cardservice.connection.getReader()}')
            try:
                self.cardservice.connection.connect(
                    protocol=SCARD_PROTOCOL_T0 | SCARD_PROTOCOL_T1,
                    mode=SCARD_SHARE_SHARED
                )
                print(f'CONNECTED TO {self.cardservice.connection.getReader()}')
                atr = self.cardservice.connection.getATR()
                print(f'ATR: {toHexString(atr)}')
            except Exception as e:
                print(f"Error connecting to card: {e}")
                self.card_label.setText("สถานะบัตรประชาชน: เชื่อมต่อบัตรไม่สำเร็จ")
                self.card_label.setStyleSheet("font-weight: bold; color: red;")
                self.disconnect_card()
                return
       
        
            # Select Thai ID card applet
            SELECT = [0x00, 0xA4, 0x04, 0x00, 0x08]
            THAI_ID_CARD = [0xA0, 0x00, 0x00, 0x00, 0x54, 0x48, 0x00, 0x01]
            response, sw1, sw2 = self.cardservice.connection.transmit(SELECT + THAI_ID_CARD)
            print(f'SELECT response: {toHexString(response)}, SW: {sw1:02x} {sw2:02x}')

            # Read data commands
            commands = {
                'cid': [0x80, 0xb0, 0x00, 0x04, 0x02, 0x00, 0x0d],
                'name_th': [0x80, 0xb0, 0x00, 0x11, 0x02, 0x00, 0x64],
                'name_en': [0x80, 0xb0, 0x00, 0x75, 0x02, 0x00, 0x64],
                'birth': [0x80, 0xb0, 0x00, 0xD9, 0x02, 0x00, 0x08],
                'gender': [0x80, 0xb0, 0x00, 0xE1, 0x02, 0x00, 0x01],
                'issuer': [0x80, 0xb0, 0x00, 0xF6, 0x02, 0x00, 0x64],
                'issue_date': [0x80, 0xb0, 0x01, 0x67, 0x02, 0x00, 0x08],
                'expire_date': [0x80, 0xb0, 0x01, 0x6F, 0x02, 0x00, 0x08],
                'address': [0x80, 0xb0, 0x15, 0x79, 0x02, 0x00, 0x64],
                # เพิ่มคำสั่งสำหรับอ่านหมายเลขคำของการมีบัตร
                'request_number': [0x80, 0xB0, 0x16, 0x19, 0x02, 0x00, 0x0E]
            }            

            # Initialize data dictionary
            self.current_data = {}

            # Read CID
            response, sw1, sw2 = self.send_apdu_with_get_response(self.cardservice.connection, commands['cid'])
            cid = self.decode_text(response)
            self.current_data['cid'] = cid
            self.cid_label.setText(f"เลขบัตรประชาชน : <b>{cid}</b>")

            # Read Thai Name
            response, sw1, sw2 = self.send_apdu_with_get_response(self.cardservice.connection, commands['name_th'])
            name_th = self.decode_text(response).replace('#', ' ')
            self.current_data['name_th'] = name_th
            self.name_th_label.setText(f"ชื่อ - นามสกุล : <b>{name_th}</b>")

            # Read English Name
            response, sw1, sw2 = self.send_apdu_with_get_response(self.cardservice.connection, commands['name_en'])
            name_en = self.decode_text(response).replace('#', ' ')
            self.current_data['name_en'] = name_en
            self.name_en_label.setText(f"ชื่อ - นามสกุล (Eng) : <b>{name_en}</b>")

            # Read Birth Date
            response, sw1, sw2 = self.send_apdu_with_get_response(self.cardservice.connection, commands['birth'])
            birth_raw = self.decode_text(response)
            birth_th, birth_en = self.parse_thai_date(birth_raw)
            self.current_data['birth_th'] = birth_th
            self.current_data['birth_en'] = birth_en
            self.birth_th_label.setText(f"วันเกิด (ไทย) : <b>{birth_th}</b>")
            self.birth_en_label.setText(f"วันเกิด (Eng) : <b>{birth_en}</b>")

            # Read Gender
            response, sw1, sw2 = self.send_apdu_with_get_response(self.cardservice.connection, commands['gender'])
            gender_code = self.decode_text(response)
            gender = "ชาย" if gender_code == "1" else "หญิง" if gender_code == "2" else gender_code
            self.current_data['gender'] = gender
            self.gender_label.setText(f"เพศ : <b>{gender}</b>")

            # Read Issue Date
            response, sw1, sw2 = self.send_apdu_with_get_response(self.cardservice.connection, commands['issue_date'])
            issue_raw = self.decode_text(response)
            issue_th, issue_en = self.parse_thai_date(issue_raw)
            self.current_data['issue_date'] = issue_th
            self.current_data['issue_date_en'] = issue_en
            self.issue_date_label.setText(f"วันที่ออกบัตร : <b>{issue_th}</b>")
            self.issue_date_en_label.setText(f"วันที่ออกบัตร (Eng) : <b>{issue_en}</b>")

            # Read Expire Date
            response, sw1, sw2 = self.send_apdu_with_get_response(self.cardservice.connection, commands['expire_date'])
            expire_raw = self.decode_text(response)
            expire_th, expire_en = self.parse_thai_date(expire_raw)
            self.current_data['expire_date'] = expire_th
            self.current_data['expire_date_en'] = expire_en
            self.expire_date_label.setText(f"วันที่หมดอายุ : <b>{expire_th}</b>")
            self.expire_date_en_label.setText(f"วันที่หมดอายุ (Eng) : <b>{expire_en}</b>")
            
            # Read Issuer (สถานที่ออกบัตร)
            response, sw1, sw2 = self.send_apdu_with_get_response(self.cardservice.connection, commands['issuer'])
            issuer_name = self.decode_text(response)
            self.current_data['issuer'] = issuer_name
            self.issuer_label.setText(f"สถานที่ออกบัตร : <b>{issuer_name}</b>")
            
            # Read Address/Issuer
            response, sw1, sw2 = self.send_apdu_with_get_response(self.cardservice.connection, commands['address'])
            address = self.decode_text(response).replace('#', ' ')
            self.current_data['address'] = address
            self.address_label.setText(f"ที่อยู่ : <b>{address}</b>")
            
            # Read Request Number (หมายเลขคำของการมีบัตร)
            response, sw1, sw2 = self.send_apdu_with_get_response(self.cardservice.connection, commands['request_number'])
            request_number = self.decode_text(response)
            self.current_data['request_number'] = request_number
            self.request_number_label.setText(f"<b>{request_number}</b>")

            # Read Photo
            photo_data = b''
            for i in range(20):
                cmd = [0x80, 0xB0, 0x01 + i, 0x7B - i, 0x02, 0x00, 0xFF]
                try:
                    response, sw1, sw2 = self.send_apdu_with_get_response(self.cardservice.connection, cmd)
                    if sw1 == 0x90 and sw2 == 0x00:
                        photo_data += bytes(response)
                    else:
                        break
                except Exception as e:
                    print(f"Photo read error: {e}")
                    break

            if photo_data and len(photo_data) > 100:
                try:
                    pixmap = QPixmap()
                    if pixmap.loadFromData(photo_data):
                        scaled_pixmap = pixmap.scaled(
                            self.photo_label.size(), 
                            Qt.KeepAspectRatio, 
                            Qt.SmoothTransformation
                        )
                        self.photo_label.setPixmap(scaled_pixmap)
                        self.current_data['photo'] = photo_data
                        print("Photo loaded successfully")
                    else:
                        print("Failed to load photo from data")
                except Exception as e:
                    print(f"Photo display error: {e}")

            # Disconnect from card after reading
            self.disconnect_card()
            
            # Update status
            self.card_label.setText("สถานะบัตรประชาชน: อ่านข้อมูลสำเร็จ")
            self.card_label.setStyleSheet("font-weight: bold; color: green;")
            self.save_button.setEnabled(False) 
            
            # # เปิดใช้งานปุ่มสแกนใบหน้า
            if 'photo' in self.current_data:
                self.scan_face_button.setEnabled(True)
                print("Face scan button enabled")           
            
            
            
        except NoCardException:
            self.card_label.setText("สถานะบัตรประชาชน: ไม่พบบัตร ภายในเวลาที่กำหนด")
            self.card_label.setStyleSheet("font-weight: bold; color: red;")
            self.scan_face_button.setEnabled(False)
            self.disconnect_card()
        except Exception as e:
            print(f"Error reading ID card: {e}")
            self.card_label.setText("สถานะบัตรประชาชน: เกิดข้อผิดพลาดในการอ่านข้อมูล")
            self.card_label.setStyleSheet("font-weight: bold; color: red;")
            self.scan_face_button.setEnabled(False)
            self.disconnect_card()        

        # ------------------- Save Data -------------------

    def save_data(self):
            """บันทึกข้อมูลตามตัวเลือกที่เลือก (Text, Excel, Image, SQLite)
            *เงื่อนไข: ต้องมีสถานะใบหน้าตรงกับบัตรเท่านั้นจึงจะบันทึก*"""


            # ตรวจสอบว่ามีข้อมูลหรือไม่
            if not self.current_data or not self.current_data.get('cid'):
                QMessageBox.warning(self, "บันทึกข้อมูล", "ไม่พบข้อมูลบัตรประชาชนที่ต้องการบันทึก (CID ว่างเปล่า)")
                return

            # ---------------- 1. กำหนดสถานะยืนยันใบหน้า (Face Verification Status) ---------------- 
            print("DEBUG: เริ่มต้นการบันทึกข้อมูล...'{self.face_verify_status.text()}'")
            face_status_text = self.face_verify_status.text()
            print(f"DEBUG: สถานะใบหน้าที่อ่านได้: '{face_status_text}'") 
            # แปลงสถานะเป็นตัวเลข (0=ไม่ตรวจสอบ, 1=ตรงกับบัตร, 2=ไม่ตรงกับบัตร)
            face_verified_status = 0 

            # ใช้ข้อความที่ตรงตัวทุกตัวอักษร
            if face_status_text == "ผลการตรวจสอบ: ตรงกับบัตรประชาชน":
                face_verified_status = 1
                face_status_text = "ตรงกับบัตรประชาชน"
            elif face_status_text == "ผลการตรวจสอบ: ไม่ตรงกับบัตรประชาชน":
                face_verified_status = 2
                face_status_text = "ไม่ตรงกับบัตรประชาชน"
          
            
            # ---------------- (ดำเนินการบันทึกต่อเมื่อเงื่อนไขผ่านเท่านั้น) ----------------
            
            # สร้างโครงสร้าง Folder
            base_folder = self.folder_path.text()
            main_data_folder = os.path.join(base_folder, "ID_Card_Data")
            if not os.path.exists(main_data_folder):
                os.makedirs(main_data_folder)
                
            today = datetime.now()
            date_str = today.strftime("%d%m%Y")
            base_filename = f"{date_str}_ID_Card"
            
            # ---------------- 3. บันทึกรูปภาพ (Image) และเตรียม Path ----------------
            photo_filepath = ""
            image_success = False
            if self.image_cb.isChecked() and 'photo' in self.current_data:
                # ใช้โครงสร้างโฟลเดอร์จากโค้ดเดิมของคุณ: /ID_Card_Data/Photo/DDMMYYYY/
                photo_folder = os.path.join(main_data_folder, "Photo")
                if not os.path.exists(photo_folder):
                    os.makedirs(photo_folder)
                
                date_folder = os.path.join(photo_folder, date_str)
                if not os.path.exists(date_folder):
                    os.makedirs(date_folder)
                
                timestamp = datetime.now().strftime("%H%M%S")
                cid = self.current_data.get('cid', 'unknown')
                filename = f"{cid}_{timestamp}.jpg"
                photo_filepath = os.path.join(date_folder, filename)
                
                try:
                    with open(photo_filepath, 'wb') as f:
                        f.write(self.current_data['photo'])
                    print(f"Photo saved: {photo_filepath}")
                    image_success = True
                except Exception as e:
                    print(f"Error saving image: {e}")
                    QMessageBox.critical(self, "Image Save Error", f"ไม่สามารถบันทึกไฟล์รูปภาพได้: {e}")
                    
            # เก็บ path รูปภาพชั่วคราวเพื่อให้ save_to_sqlite และ Text File นำไปใช้
            self.current_data['photo_path'] = photo_filepath

            # ---------------- 4. บันทึกข้อมูลลง SQLite Database ----------------
            sqlite_success = False
            if self.sqlite_cb.isChecked():
                sqlite_success = self.save_to_sqlite(self.current_data, face_verified_status)
            
            if sqlite_success:
                # นับจำนวนครั้งที่สแกนของ CID นี้
                try:
                    conn = sqlite3.connect(self.db_name)
                    cursor = conn.cursor()
                    cursor.execute("SELECT COUNT(*) FROM id_card_records WHERE cid = ?", 
                                (self.current_data['cid'],))
                    scan_count = cursor.fetchone()[0]
                    conn.close()
                    print(f"📊 CID {self.current_data['cid']} ถูกสแกนไปแล้ว {scan_count} ครั้ง")
                except:
                    pass            
            # ---------------- 5. บันทึก Text File ----------------
            text_success = False
            if self.text_file_cb.isChecked():
                filename = f"{base_filename}.txt"
                filepath = os.path.join(main_data_folder, filename)
                file_exists = os.path.exists(filepath)
                try:
                    with open(filepath, 'a', encoding='utf-8') as f:
                        if file_exists:
                            f.write("\n" + "=" * 50 + "\n")
                        else:
                            f.write("Thai ID Card Data Log\n")
                            f.write("=" * 50 + "\n")
                            
                        f.write(f"Record Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                        f.write(f"Citizen ID: {self.current_data.get('cid', '')}\n")
                        f.write(f"Thai Name: {self.current_data.get('name_th', '')}\n")
                        f.write(f"English Name: {self.current_data.get('name_en', '')}\n")
                        f.write(f"Gender: {self.current_data.get('gender', '')}\n")
                        f.write(f"Birth Date (Thai): {self.current_data.get('birth_th', '')}\n")
                        f.write(f"Birth Date (Eng): {self.current_data.get('birth_en', '')}\n")
                        f.write(f"Issuer: {self.current_data.get('issuer', '')}\n")
                        f.write(f"Issue Date: {self.current_data.get('issue_date', '')}\n")
                        f.write(f"Expire Date: {self.current_data.get('expire_date', '')}\n")
                        f.write(f"Address: {self.current_data.get('address', '')}\n")
                        f.write(f"Request Number: {self.current_data.get('request_number', '')}\n")
                        f.write(f"Face Verify Status: {face_status_text}\n")
                        f.write(f"Photo Path: {photo_filepath}\n")
                    text_success = True
                except Exception as e:
                    print(f"Error saving text file: {e}")
                    QMessageBox.critical(self, "Text File Save Error", f"ไม่สามารถบันทึก Text File ได้: {e}")
                    
            # ---------------- 6. บันทึก Excel File ----------------
            excel_success = False
            if self.excel_file_cb.isChecked() and OPENPYXL_AVAILABLE:
                filename = f"{base_filename}.xlsx"
                filepath = os.path.join(main_data_folder, filename)
                
                try:
                    # ฟังก์ชันแปลงวันที่ (ใช้โค้ดเดิมของคุณ)
                    def convert_to_short_date(full_date_str):
                        if not full_date_str or full_date_str in ("ไม่ระบุ", "ตลอดชีพ"):
                            return full_date_str
                        try:
                            month_map = {
                                'มกราคม': '1', 'กุมภาพันธ์': '2', 'มีนาคม': '3', 'เมษายน': '4',
                                'พฤษภาคม': '5', 'มิถุนายน': '6', 'กรกฎาคม': '7', 'สิงหาคม': '8',
                                'กันยายน': '9', 'ตุลาคม': '10', 'พฤศจิกายน': '11', 'ธันวาคม': '12'
                            }
                            parts = full_date_str.split()
                            if len(parts) >= 3 and parts[1] in month_map:
                                return f"{parts[0]}/{month_map[parts[1]]}/{parts[2]}"
                        except:
                            pass
                        return full_date_str

                    # ตรวจสอบว่าไฟล์ Excel เปิดอยู่หรือไม่ (ใช้โค้ดเดิมของคุณ)
                    if os.path.exists(filepath):
                        try:
                            with open(filepath, 'r+b') as test_file:
                                pass 
                        except (PermissionError, IOError):
                            reply = QMessageBox.question(self, "ไฟล์ Excel กำลังใช้งานอยู่", 
                                f"ไฟล์ {filename} กำลังถูกใช้งานอยู่ (อาจเปิดอยู่ใน Excel)\n"
                                "คลิก 'Retry' เพื่อลองใหม่หลังจากปิดไฟล์แล้ว\n"
                                "คลิก 'Cancel' เพื่อข้าม",
                                QMessageBox.Retry | QMessageBox.Cancel)
                            if reply == QMessageBox.Retry:
                                try:
                                    with open(filepath, 'r+b') as test_file: pass
                                except (PermissionError, IOError):
                                    QMessageBox.warning(self, "ไฟล์ยังใช้งานอยู่", f"ไฟล์ {filename} ยังถูกใช้งานอยู่\nกรุณาปิด Excel และลองใหม่อีกครั้ง")
                                    return 
                            else:
                                return 
                    
                    if os.path.exists(filepath):
                        workbook = load_workbook(filepath)
                        worksheet = workbook.active
                    else:
                        workbook = Workbook()
                        worksheet = workbook.active
                        worksheet.title = "ID Card Data"

                    # หาแถวสุดท้ายที่มีข้อมูลและคำนวณตำแหน่งใหม่ (ใช้โค้ดเดิมของคุณ)
                    last_row = 1
                    for row in range(1, worksheet.max_row + 1):
                        if worksheet[f'A{row}'].value and str(worksheet[f'A{row}'].value).startswith('No.'):
                            last_row = max(last_row, row)

                    if last_row == 1 and not worksheet['A1'].value:
                        new_entry_number = 1
                        start_row = 1
                    else:
                        last_entry_value = worksheet[f'A{last_row}'].value
                        new_entry_number = int(str(last_entry_value).replace('No.', '')) + 1 if last_entry_value and str(last_entry_value).startswith('No.') else 1
                        start_row = last_row + 8 

                    # แปลงวันที่เป็นรูปแบบสั้น
                    birth_short = convert_to_short_date(self.current_data.get('birth_th', ''))
                    issue_short = convert_to_short_date(self.current_data.get('issue_date', ''))
                    expire_short = convert_to_short_date(self.current_data.get('expire_date', ''))

                    # ตั้งค่าความสูงของแถว
                    for i in range(7):
                        worksheet.row_dimensions[start_row + i].height = 16 

                    data_mapping = {
                        f'A{start_row}': f'No.{new_entry_number}',
                        f'A{start_row + 1}': 'เลขบัตรประชาชน', 
                        f'C{start_row + 1}': 'ชื่อ', 
                        f'D{start_row + 1}': self.current_data.get('name_th', ''),
                        f'A{start_row + 2}': f"{self.current_data.get('cid', '')}", 
                        f'C{start_row + 2}': 'Name', 
                        f'D{start_row + 2}': self.current_data.get('name_en', ''),
                        f'B{start_row + 3}': 'เกิดวันที่', 
                        f'C{start_row + 3}': birth_short,
                        f'B{start_row + 4}': 'Address', 
                        f'C{start_row + 4}': self.current_data.get('address', ''),
                        f'A{start_row + 5}': 'ออกบัตรเมื่อ', 
                        f'B{start_row + 5}': 'บัตรหมดอายุ', 
                        f'C{start_row + 5}': 'สถานที่ออกบัตร',
                        f'A{start_row + 6}': issue_short, 
                        f'B{start_row + 6}': expire_short, 
                        f'C{start_row + 6}': self.current_data.get('issuer', ''),
                        f'D{start_row + 3}': 'สถานะใบหน้า:',
                        f'E{start_row + 3}': face_status_text,
                        
                        # ✅ เพิ่ม label สำหรับรูปภาพ
                        f'F{start_row}': 'รูปจากบัตร',  # Header คอลัมน์ F
                        f'H{start_row}': 'รูปจากกล้อง',  # Header คอลัมน์ H
                    }

                    # กำหนด font styles และใส่ข้อมูลลงใน cells (ใช้โค้ดเดิมของคุณ)
                    title_font = Font(name='Aptos Narrow', bold=True, size=11)
                    label_font = Font(name='Aptos Narrow', size=11)
                    data_font = Font(name='TAptos Narrow', size=11)
                    id_font = Font(name='Aptos Narrow', size=11)

                    for cell_address, value in data_mapping.items():
                        worksheet[cell_address] = value
                        
                        if cell_address.startswith('A') and str(value).startswith('No.'):
                            worksheet[cell_address].font = title_font
                        elif cell_address == f'A{start_row + 2}':
                            worksheet[cell_address].font = id_font
                        elif any(label in cell_address for label in [f'A{start_row + 1}', f'C{start_row + 1}', f'C{start_row + 2}', f'B{start_row + 3}', f'B{start_row + 4}', f'A{start_row + 5}', f'B{start_row + 5}', f'C{start_row + 5}', f'D{start_row + 3}', f'D{start_row + 4}']):
                            worksheet[cell_address].font = label_font 
                        else:
                            worksheet[cell_address].font = data_font 
                            
                        worksheet[cell_address].alignment = Alignment(vertical='center')
                        
                    # การเพิ่มรูปภาพ
                    if 'photo' in self.current_data:
                        try:
                            image_stream = io.BytesIO(self.current_data['photo'])
                            img = OpenpyxlImage(image_stream)
                            img.width = 110 ; 
                            img.height = 130
                            worksheet.add_image(img, f'F{start_row + 1}')

                            if 'captured_face_photo' in self.current_data: 
                                capture_image_stream = io.BytesIO(self.current_data['captured_face_photo'])
                                capture_img = OpenpyxlImage(capture_image_stream)
                                capture_img.width = 110 ; capture_img.height = 130 
                                worksheet.add_image(capture_img, f'H{start_row + 1}')
                        except Exception as e:
                            print(f"ผิดพลาดในการเพิ่มรูปภาพใน Excel: {e}")

                    # ปรับความกว้างของคอลัมน์
                    worksheet.column_dimensions['A'].width = 16.25 
                    worksheet.column_dimensions['B'].width = 12.33 
                    worksheet.column_dimensions['C'].width = 48.67
                    worksheet.column_dimensions['D'].width = 16.00
                    worksheet.column_dimensions['E'].width = 23.22
                    worksheet.column_dimensions['F'].width = 14.22
                    worksheet.column_dimensions['G'].width = 3 
                    worksheet.column_dimensions['H'].width = 18.00
                    
                    # บันทึกไฟล์
                    workbook.save(filepath)
                    excel_success = True
                    
                except Exception as e:
                    excel_success = False
                    if "Permission denied" in str(e) or "being used by another process" in str(e):
                        QMessageBox.critical(self, "ไม่สามารถบันทึกได้", f"ไม่สามารถบันทึกไฟล์ {filename} ได้\nไฟล์อาจถูกเปิดใช้งานอยู่ใน Excel\nกรุณาปิด Excel และลองใหม่อีกครั้ง")
                    else:
                        QMessageBox.critical(self, "เกิดข้อผิดพลาด", f"ไม่สามารถบันทึกไฟล์ Excel ได้: {e}")
                    
            # ---------------- 7. การทำความสะอาดและแจ้งเตือนสรุป ----------------
            
            # ล้าง path รูปภาพชั่วคราวออกจาก current_data
            if 'photo_path' in self.current_data:
                del self.current_data['photo_path']

            # สรุปสถานะการบันทึก
            save_status_msg = []
            if self.text_file_cb.isChecked():
                save_status_msg.append(f"Text file: {'บันทึกสำเร็จ' if text_success else 'ไม่สำเร็จ'}")
            if self.excel_file_cb.isChecked():
                save_status_msg.append(f"Excel file: {'บันทึกสำเร็จ' if excel_success else 'ไม่สำเร็จ'}")
            if self.image_cb.isChecked():
                status = f"บันทึกสำเร็จ ({os.path.basename(photo_filepath)})" if image_success else "ไม่บันทึก/ผิดพลาด"
                save_status_msg.append(f"Image: {status}")
            if self.sqlite_cb.isChecked():
                save_status_msg.append(f"SQLite DB: {'บันทึกสำเร็จ/อัปเดต' if sqlite_success else 'ผิดพลาด'}")
            
            if save_status_msg:
                self.save_button.setEnabled(False)
                QMessageBox.information(self, "บันทึกข้อมูลสำเร็จ", 
                                        f"บันทึกข้อมูลของ CID **{self.current_data['cid']}** เรียบร้อยแล้ว:\n\n"
                                        + "\n".join(save_status_msg))
            else:
                QMessageBox.information(self, "บันทึกข้อมูล", "ไม่มีตัวเลือกการบันทึกใดถูกเลือก")

    # ------------------- Clear & Exit -------------------
    def clear_data(self):
        """Clear all displayed data and photo"""
        self.cid_label.setText("เลขประจำตัวประชาชน :")
        self.name_th_label.setText("ชื่อ - นามสกุล :")
        self.name_en_label.setText("ชื่อ - นามสกุล (Eng) :")
        self.birth_th_label.setText("วันเกิด (ไทย) :")
        self.birth_en_label.setText("วันเกิด (Eng) :")
        self.gender_label.setText("เพศ :")
        self.issue_date_label.setText("วันที่ออกบัตร :")
        self.issue_date_en_label.setText("วันที่ออกบัตร (Eng) :")
        self.expire_date_label.setText("วันที่หมดอายุ :")
        self.expire_date_en_label.setText("วันที่หมดอายุ (Eng) :")
        self.issuer_label.setText("สถานที่ออกบัตร :")
        self.address_label.setText("ที่อยู่ :")
        self.photo_label.clear()
        self.photo_label.setText("รูปภาพจากบัตรประชาชน")
        self.request_number_label.setText("หมายเลขคำขอการมีบัตร :")
        
        # ✅ รีเซ็ตข้อมูลอย่างสมบูรณ์
        self.current_data = {}
        
        # รีเซ็ตสถานะใบหน้า
        self.face_verify_status.setText("รอการตรวจสอบ")
        self.face_verify_status.setStyleSheet("""
            font-size: 14px;
            font-weight: bold;
            color: #6B7280;
            padding: 10px;
        """)
        self.face_progress.setVisible(False)
        self.face_progress.setValue(0)
        
        # ✅ รีเซ็ตตัวแปรที่เกี่ยวข้อง
        self.captured_face_frame = None
        self.is_face_scan_mode = False
        
        # ปิดปุ่มที่ไม่ควรใช้งานได้
        self.save_button.setEnabled(False)
        self.scan_face_button.setEnabled(False)
        
        print("Data cleared")
        
    def cleanup_resources(self):
        """Complete cleanup of all resources"""
        print("Starting cleanup...")
        
        # Stop timer
        try:
            if hasattr(self, 'timer') and self.timer.isActive():
                self.timer.stop()
                print("Timer stopped")
        except Exception as e:
            print(f"Error stopping timer: {e}")
        
        # Disconnect card
        try:
            self.disconnect_card()
        except Exception as e:
            print(f"Error disconnecting card: {e}")
        
        # Remove card observer
        try:
            if hasattr(self, 'cardmonitor') and hasattr(self, 'cardobserver'):
                self.cardmonitor.deleteObserver(self.cardobserver)
                print("Card observer removed")
        except Exception as e:
            print(f"Error removing card observer: {e}")
        
        print("Cleanup completed")

    def exit_application(self):
        self.disconnect_card()
        self.cardmonitor.deleteObserver(self.cardobserver)
        QApplication.quit()
        
    def closeEvent(self, event):
        """Clean up when closing the application"""
        try:
            self.cleanup_resources()
            print("Application closed successfully")
            event.accept()
            
        except Exception as e:
            print(f"Error during close: {e}")
            event.accept()

    def start_webcam(self):
        """เริ่มต้นการทำงานของกล้องเว็บแคม"""
        if self.camera is None:
            self.camera = cv2.VideoCapture(0)  # เปิดกล้องตัวที่ 0 (กล้องหลัก)
            self.webcam_running = True
        if self.camera.isOpened():
            self.webcam_timer.start(30)
            if self.webcam_label is not None:
               self.webcam_label.setText("กำลังเปิดกล้อง...")            
            print("Webcam started.")
        else:
            if self.webcam_label is not None:
                self.webcam_label.setText("ไม่พบกล้อง")
            QMessageBox.critical(self, "ข้อผิดพลาด", "ไม่สามารถเปิดกล้องเว็บแคมได้")
            self.camera = None
            self.webcam_running = False # อัปเดตสถานะเมื่อกล้องเปิดไม่สำเร็จ

    def stop_webcam(self):
        """หยุดการทำงานของกล้องเว็บแคม"""
        if self.webcam_timer.isActive():
            self.webcam_timer.stop()
        
        if self.camera is not None:
            self.camera.release()
            self.camera = None
            print("Webcam stopped.")
            
        # เคลียร์หน้าจอแสดงผล
        self.webcam_label.setText("รูปภาพ")
        self.webcam_label.setPixmap(QPixmap()) # ลบภาพออกจาก Label

    # ส่วนที่ 2: แก้ไข scan_and_compare_face() - สลับโหมดกล้อง
    def scan_and_compare_face(self):
        """เปิด/ปิดโหมดสแกนใบหน้าใน webcam_label"""
        
        # ✅ ตรวจสอบข้อมูลบัตรก่อน
        if not self.current_data:
            QMessageBox.warning(self, "ไม่พบข้อมูลบัตร", 
                "กรุณาอ่านข้อมูลบัตรประชาชนก่อน")
            return
            
        if 'photo' not in self.current_data:
            QMessageBox.warning(self, "ไม่พบรูปบัตร", 
                "ไม่พบรูปภาพในบัตรประชาชน\nกรุณาอ่านบัตรใหม่อีกครั้ง")
            return
        
        if self.camera is None:
            # เปิดกล้อง
            self.camera = cv2.VideoCapture(0)
            if not self.camera.isOpened():
                QMessageBox.critical(self, "ข้อผิดพลาด", 
                    "ไม่สามารถเปิดกล้องได้\nกรุณาตรวจสอบว่าใช้งานกล้องไม่ได้")
                self.camera = None
                return
            
            # ตั้งค่าความละเอียด
            self.camera.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
            self.camera.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)
            
            # เริ่มโหมดสแกนใบหน้า
            self.is_face_scan_mode = True
            self.captured_face_frame = None
            self.webcam_timer.start(30)  # เปลี่ยนจาก 1 เป็น 30 milliseconds
            
            # เปลี่ยนข้อความปุ่ม
            self.scan_face_button.setText("Capture")
            print("เปิดโหมดสแกนใบหน้า - คลิกปุ่ม 'ถ่ายรูป' เพื่อถ่าย")
            
        elif not self.is_face_scan_mode:
            # เข้าโหมดสแกนใบหน้า
            self.is_face_scan_mode = True
            self.captured_face_frame = None
            self.scan_face_button.setText("Capture")
            print("เข้าโหมดสแกนใบหน้า")
            
        else:
            # ถ่ายรูปและเปรียบเทียบ
            if self.captured_face_frame is not None:
                # หยุดกล้องชั่วคราว
                self.is_face_scan_mode = False
                
                print("กำลังประมวลผลและเปรียบเทียบใบหน้า...")
                
                # ✅ เปรียบเทียบใบหน้า
                self.compare_faces(self.captured_face_frame, self.face_cascade)
                
                # รีเซ็ต
                self.captured_face_frame = None
                self.scan_face_button.setText("Face scan")
            else:
                QMessageBox.warning(self, "ไม่พบใบหน้า", 
                    "ไม่พบใบหน้าในภาพ\nกรุณาหันหน้าเข้ากล้องแล้วลองอีกครั้ง")
                
    def put_thai_text(image_bgr, text, position=(10, 30), font_size=28, color=(0, 255, 0)):
        """ฟังก์ชันสำหรับใส่ข้อความภาษาไทยลงในภาพ OpenCV ด้วย PIL"""
        # แปลง OpenCV image (BGR) -> PIL (RGB)
        img_pil = Image.fromarray(cv2.cvtColor(image_bgr, cv2.COLOR_BGR2RGB))
        draw = ImageDraw.Draw(img_pil)

        # โหลดฟอนต์ไทย (คุณต้องมีไฟล์ฟอนต์ในเครื่อง เช่น THSarabunNew.ttf)
        font = ImageFont.truetype("THSarabunNew.ttf", font_size)

        # เขียนข้อความ
        draw.text(position, text, font=font, fill=color)

        # แปลงกลับเป็น OpenCV (BGR)
        return cv2.cvtColor(np.array(img_pil), cv2.COLOR_RGB2BGR)
    
    # ส่วนที่ 3: แก้ไข update_frame() - ใช้ Haar Cascade สำหรับ real-time (เร็ว)
    def update_frame(self):
        """ดึงเฟรมภาพจากกล้องและแสดงผลใน webcam_label"""
        if self.camera is None or not self.camera.isOpened():
            return

        ret, frame = self.camera.read()
        if not ret:
            return
        
        display_frame = frame.copy()
        
        if self.is_face_scan_mode:
            
            if self.face_cascade is None or self.face_cascade.empty():
                print("Face detection disabled: Haar Cascade model not loaded correctly.")
                self.is_face_scan_mode = False
                return
                
            try:
                gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
                
                faces = self.face_cascade.detectMultiScale(
                    gray, 
                    scaleFactor=1.3, 
                    minNeighbors=5,
                    minSize=(30, 30)
                )
                
                for (x, y, w, h) in faces:
                    cv2.rectangle(display_frame, (x, y), (x+w, y+h), (0, 255, 0), 2)
                    cv2.putText(display_frame, "Face Detected", (x, y-10),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2)
                
                if len(faces) > 0:
                    cv2.putText(display_frame, "Click 'Capture' button to take photo", (10, 30),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2)
                    self.captured_face_frame = frame.copy()
                else:
                    cv2.putText(display_frame, "No face detected - Please face camera", (10, 30),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 0, 255), 2)
                    self.captured_face_frame = None
                    
            except Exception as e:
                print(f"Face detection error: {e}")
        
        frame_rgb = cv2.cvtColor(display_frame, cv2.COLOR_BGR2RGB)
        h, w, ch = frame_rgb.shape
        bytes_per_line = ch * w
        qt_image = QImage(frame_rgb.data, w, h, bytes_per_line, QImage.Format_RGB888)
        pixmap = QPixmap.fromImage(qt_image)
        
        pixmap = pixmap.scaled(
            self.webcam_label.size(), 
            Qt.KeepAspectRatio, 
            Qt.SmoothTransformation
        )
        
        self.webcam_label.setPixmap(pixmap)
 

    def init_db(self):
        """เชื่อมต่อฐานข้อมูลและสร้างตารางหากยังไม่มี"""
        try:
            conn = sqlite3.connect(self.db_name)
            cursor = conn.cursor()
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS id_card_records (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    record_timestamp TEXT NOT NULL,
                    scan_timestamp TEXT,
                    cid TEXT NOT NULL,
                    name_th TEXT,
                    name_en TEXT,
                    gender TEXT,
                    birth_date TEXT,
                    issue_date TEXT,
                    expire_date TEXT,
                    issuer TEXT,
                    address TEXT,
                    request_number TEXT,
                    card_photo BLOB,                    
                    captured_face_photo BLOB,           
                    photo_saved_path TEXT,              
                    face_verified INTEGER DEFAULT 0
                )
            """)
            conn.commit()
            conn.close()
            print(f"SQLite Database initialized at: {self.db_name}")
        except Exception as e:
            QMessageBox.critical(self, "SQLite Error", 
                                f"ไม่สามารถเริ่มต้นฐานข้อมูล SQLite ได้: {e}")

    def save_to_sqlite(self, data, face_verified_status):
        """บันทึกข้อมูลบัตรประชาชนลงในฐานข้อมูล SQLite พร้อมรูปภาพ - บันทึกทุกครั้ง"""
        if not data.get('cid'):
            return False

        try:
            conn = sqlite3.connect(self.db_name)
            cursor = conn.cursor()
            
            # ✅ แปลงรูปภาพจากบัตรเป็น binary (BLOB)
            card_photo_blob = data.get('photo', None)
            
            # ✅ แปลงรูปภาพจากกล้องเป็น binary (BLOB)
            captured_photo_blob = data.get('captured_face_photo', None)
            
            # เก็บ path ไว้ด้วย (optional)
            photo_path = data.get('photo_path', '')
            
            # บันทึกเวลาปัจจุบันทุกครั้ง
            record_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            scan_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            record = (
                record_time,
                scan_time,  # ✅ เพิ่มเวลาสแกน
                data.get('cid', ''),
                data.get('name_th', ''),
                data.get('name_en', ''),
                data.get('gender', ''),
                data.get('birth_th', ''), 
                data.get('issue_date', ''),
                data.get('expire_date', ''),
                data.get('issuer', ''),
                data.get('address', ''),
                data.get('request_number', ''),
                card_photo_blob,
                captured_photo_blob,
                photo_path,
                face_verified_status
            )

            # ✅ เปลี่ยนจาก INSERT OR REPLACE เป็น INSERT เพื่อบันทึกทุกครั้ง
            cursor.execute("""
                INSERT INTO id_card_records (
                    record_timestamp, scan_timestamp, cid, name_th, name_en, gender, birth_date, 
                    issue_date, expire_date, issuer, address, request_number, 
                    card_photo, captured_face_photo, photo_saved_path, face_verified
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, record)
            
            conn.commit()
            conn.close()
            
            print(f"✅ บันทึกข้อมูล CID: {data.get('cid')} ครั้งที่ {cursor.lastrowid} เวลา: {scan_time}")
            return True
            
        except Exception as e:
            print(f"SQLite save error: {e}")
            QMessageBox.critical(self, "SQLite Error", 
                                f"ไม่สามารถบันทึกข้อมูลลงฐานข้อมูลได้: {e}")
            return False
    
# ✅ วาง resizeEvent และ update_fonts ตรงนี้ (นอก __init__)
    def resizeEvent(self, event):
        """ปรับ font size เมื่อมีการ resize หน้าต่าง"""
        super().resizeEvent(event)
        
        if not self.ui_ready:
            return
        
        current_width = self.width()
        current_height = self.height()
        
        width_scale = current_width / self.base_width
        height_scale = current_height / self.base_height
        scale_factor = min(width_scale, height_scale)
        
        new_font_size = max(10, int(self.base_font_size * scale_factor))
        self.update_fonts(new_font_size)

    def update_fonts(self, font_size):
        """อัปเดต font size ของ widgets ทั้งหมด"""
        
        # ตรวจสอบว่า widgets ถูกสร้างแล้วหรือยัง
        if not hasattr(self, 'cid_label'):
            return
        
        # อัปเดต labels ด้วย stylesheet แทน setFont
        info_labels = [
            self.cid_label, self.name_th_label, self.name_en_label,
            self.birth_th_label, self.birth_en_label, self.gender_label,
            self.issue_date_label, self.issue_date_en_label,
            self.expire_date_label, self.expire_date_en_label,
            self.issuer_label, self.address_label, self.request_number_label
        ]
        
        for label in info_labels:
            label.setStyleSheet(f"font-family: 'Prompt'; font-size: {font_size}px; color: #333333;")
        
        # อัปเดต status labels
        status_size = max(9, int(font_size * 0.85))
        self.reader_label.setStyleSheet(f"font-family: 'Prompt'; font-size: {status_size}px; font-weight: bold;")
        # self.card_label.setStyleSheet(f"font-family: 'Prompt'; font-size: {status_size}px;")
        
 
        
        # อัปเดต GroupBox titles
        group_size = int(font_size * 1.1)
        for widget in self.findChildren(QGroupBox):
            current_style = widget.styleSheet()
            # แทนที่ font-size ใน stylesheet
            import re
            new_style = re.sub(r'font-size:\s*\d+px', f'font-size: {group_size}px', current_style)
            widget.setStyleSheet(new_style)   

    def open_database_viewer(self):
        '''เปิดหน้าต่างดูฐานข้อมูล'''
        try:
            if not os.path.exists(self.db_name):
                QMessageBox.warning(self, "ไม่พบฐานข้อมูล", 
                                "ยังไม่มีข้อมูลในฐานข้อมูล\nกรุณาอ่านบัตรและบันทึกข้อมูลก่อน")
                return
            
            self.db_viewer = DatabaseViewerWindow(db_path=self.db_name, parent=self)
            self.db_viewer.show()
            
        except Exception as e:
            QMessageBox.critical(self, "ข้อผิดพลาด", 
                            f"ไม่สามารถเปิดหน้าดูฐานข้อมูลได้:\n{str(e)}")

            
    def to_thai_num(num_str):
        mapping = {'0': '๐', '1': '๑', '2': '๒', '3': '๓', '4': '๔',
                '5': '๕', '6': '๖', '7': '๗', '8': '๘', '9': '๙'}
        result = ''.join(mapping.get(digit, digit) for digit in str(num_str))
        return result               
# ------------------- Main -------------------
if __name__ == "__main__":
    import sys
    app = QApplication(sys.argv)
    app.setFont(QFont("Prompt", 12))
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())
